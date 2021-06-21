import player
from openpyxl import load_workbook
import string

int_to_letter = dict(zip(range(1, 27), string.ascii_uppercase))


class ParseFile:

    def __init__(self, fn):
        # TODO overhaul this code to read from a CSV file rather than a spreadsheet.
        self.wb = load_workbook(fn)
        self.sheet = self.wb['data']
        self.player = player.Player("priest", 'base')
        col = 2
        self.toons = [self.parse_toon(self.player, col)]

        col = 3
        while self.check_comparison(col) is True:
            comp_num = 1

            title = self.sheet.cell(1, col).value
            if title is not None:
                comp_name = title
            else:
                comp_name = comp_num
            new_toon = player.Player("priest", comp_name)
            self.toons.append(self.parse_toon(new_toon, col, self.toons[0]))

            comp_num += 1
            col += 1

    def check_comparison(self, col):
        col = int_to_letter[col]
        cell_range = self.sheet['{}1'.format(col):'{}13'.format(col)]
        for cell in cell_range:
            if cell[0].value is not None:
                return True
        return False

    def parse_toon(self, toon, col, base_toon=None):
        # first set baseline stats
        stats_dict = {}
        talent_dict = {}
        row = 2
        txt = self.sheet.cell(row, 1).value
        talents_row = None

        while txt != '###talents':
            val = self.sheet.cell(row, col).value
            if val is not None:
                stats_dict[txt] = val
            else:
                stats_dict[txt] = base_toon.stats_dict[txt]

            row += 1
            txt = self.sheet.cell(row, 1).value
            if txt == '###talents':
                talents_row = row + 1
        toon.assign_dict_stats(stats_dict)
        for x in (talents_row, talents_row + 3):
            talent_dict[self.sheet.cell(x, 1).value] = self.sheet.cell(x, col).value
        toon.assign_talents(talent_dict)
        return toon
